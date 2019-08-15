import openpyxl
import re
import win32com.client as win32
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

#对比的参考文件
filepath = 'OS.20190722.xlsx'
#待处理文件
fixfilepath = 'bak.20190722.xlsx'

def firststep():
    # 中间字典参照
    t = {}
    wb = openpyxl.load_workbook(filepath)
    fixwxcel = openpyxl.load_workbook(fixfilepath)

    #sheet = wb.get_sheet_names()[0]

    sheet  = wb.active
    sheet2 = fixwxcel.active

    # a = sheet["A3"].fill.start_color.index
    # print(a)

    for row in range(2,sheet.max_row+1):
        t[sheet.cell(row,10).value] = sheet.cell(row,23).value
        #c[sheet.cell(row, 10).value] = sheet.cell(row,23).fill.start_color.rgb
    # print(sheet.cell(3,1).fill.start_color)
    for row in range(2,sheet2.max_row+1):
        if sheet2.cell(row,21).value == '未完成整改':
            try:
                sheet2.cell(row,23).value = t[sheet2.cell(row,10).value]
                if re.findall('不涉及',sheet2.cell(row,23).value):
                    for col in range(1,sheet2.max_column+1):
                        sheet2.cell(row,col).fill = PatternFill("solid",fgColor='FFFF00')
                elif re.findall('备案',sheet2.cell(row,23).value):
                    for col in range(1,sheet2.max_column+1):
                        sheet2.cell(row,col).fill = PatternFill("solid",fgColor='538DD5')
                elif sheet2.cell(row,23).value == None:
                    continue
                else:
                    for col in range(1,sheet2.max_column+1):
                        sheet2.cell(row,col).fill = PatternFill("solid",fgColor='FF0000')

                #sheet2.cell(row,23).fill = PatternFill("solid",fgColor=c[sheet2.cell(row,10).value])

            except KeyError:
                print("新CVE编号")
            except TypeError:
                pass
                #print("无色号")
        else:
            continue
    # sheet2['W57'].fill = PatternFill("solid",fgColor=sheet['W57'].fill.start_color)
    fixwxcel.save('test.xlsx')

filepath1 = 'OS.20190809.xlsx'
nullfile = 'OS.null.xlsx'
fixfilepath1 = 'test.xlsx'

def secondstep():
    t = {}
    t2 = {}
    wb = openpyxl.load_workbook(filepath1)
    fixwxcel = openpyxl.load_workbook(fixfilepath1)
    nullbook = openpyxl.load_workbook(nullfile)

    sheet = wb.active
    sheet2 = fixwxcel.active
    nb = nullbook.active

    for row in range(11, sheet.max_row - 2):
        t[sheet.cell(row,2).value] = sheet.cell(row,1).value
        t2[sheet.cell(row, 6).value] = sheet.cell(row, 9).value

    listResult = []
    for row in range(2,sheet2.max_row+1):
        try:
            if sheet2.cell(row, 21).value == '未完成整改':
                lineData = [t[sheet2.cell(row,5).value],sheet2.cell(row,5).value,'无','LINUX\n主机',sheet2.cell(row,11).value,sheet2.cell(row,6).value,sheet2.cell(row,8).value,sheet2.cell(row,9).value,t2[sheet2.cell(row,6).value]]
            listResult.append(lineData)
        except KeyError:
            print("备案新CVE编号")
    # print(listResult)
    for i in range(1, len(listResult) + 1):
        for j in range(1, len(listResult[0]) + 1):
            cell = nb.cell(row=i+10, column=j)
            cell.value = listResult[i-1][j-1]
            #边框
            cell.border = Border(left=Side(border_style = 'thin',color = '000000'),
                                 right=Side(border_style='thin', color='000000'),
                                 top=Side(border_style='thin', color='000000'),
                                 bottom=Side(border_style='thin', color='000000'))
            #字体大小
            cell.font = Font(size = 10)
            #居中、自动换行
            if j == len(listResult[0]):
                cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
            else:
                cell.alignment = Alignment(horizontal='center',vertical='center',wrapText=True)

    nullbook.save('result.xlsx')

if __name__ == '__main__':
    firststep()
    secondstep()