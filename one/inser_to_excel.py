# 接受两个整数（N,M)，在第N+1行，插入M个空行，存入原文件。
import openpyxl as xls


def blankRowInsert(N, M):
    fileName = "c:\\pythonDocu\\card\\test.xlsx"
    wb = xls.load_workbook(fileName)
    sheet = wb.get_sheet_by_name('test')
    myList = dumpDataToList(sheet)
    insertLine(myList, N, M, sheet.max_column)
    setValue(sheet, myList)
    wb.save(fileName)
    wb.close()


def dumpDataToList(sheet):
    """把一个表格中的数据全部导出到一个列表"""
    listResult = []
    for i in range(1, sheet.max_row + 1):
        lineData = []
        for j in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=i, column=j)
            lineData.append(cell.value)
        listResult.append(lineData)
    return listResult


def insertLine(aList, N, M, maxColumn):
    """对列表进行添加操作操作"""
    for _ in range(1, M + 1):
        aList.insert(N, [''] * maxColumn)
    """['']*N是创建一个个数为N的空格列表，插入列表aList"""


def setValue(sheet, list):
    """把数据写回sheet"""
    for i in range(1, len(list) + 1):
        for j in range(1, len(list[0]) + 1):
            cell = sheet.cell(row=i, column=j)
            cell.value = list[i - 1][j - 1]


blankRowInsert(5, 3)  # 在第5行新增3行数据